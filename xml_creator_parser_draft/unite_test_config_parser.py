#! /usr/bin/env python
# -*- coding: utf-8 -*-

import unittest
from openpyxl import load_workbook

import config_parser

CONFIG_FILE = R'test_data/test_data_cfg.xml'
PRODUCT_OWNER_LIST_FILE = R'test_data/test_data_product_owners.xlsx'

class TestConfigParser(unittest.TestCase):
    def test_project_name_list_parser(self):
        '''get project name list from excel & xml, compare them
        '''
        cfg_parser = config_parser.CfgParser(CONFIG_FILE)
        project_list_from_xml = cfg_parser.getProjectList()
        
        wb = load_workbook(PRODUCT_OWNER_LIST_FILE)
        sheet_ranges = wb['Sheet1']
        column_length = int(sheet_ranges.dimensions.split(':')[1].split('C')[1]) 
        project_list_from_excel = []
        for i in range(2, column_length+1):
            index = str(i)
            project_list_from_excel.append(sheet_ranges['A'+index].value)
        project_list_from_excel = list(set(project_list_from_excel))

        self.assertEqual(sorted(project_list_from_xml), sorted(project_list_from_excel)) 
        
    def test_project_owner_email_parser(self):
        '''get project owner email from excel & xml, compare them
        '''
        cfg_parser = config_parser.CfgParser(CONFIG_FILE)
        project_list_from_xml = cfg_parser.getProjectList()
        for project in project_list_from_xml:
            if(project.split('-')[0]!='XenDesktop'):
                owner_email_from_xml = cfg_parser.getOwnerEmailsOfProject(project)[0]
                
                wb = load_workbook(PRODUCT_OWNER_LIST_FILE)
                sheet_ranges = wb['Sheet1']
                column_length = int(sheet_ranges.dimensions.split(':')[1].split('C')[1]) 
                for i in range(2, column_length+1):
                    index = str(i)
                    if(project==sheet_ranges['A'+index].value):
                        owner_email_from_excel = sheet_ranges['C'+index].value
                        self.assertEqual(owner_email_from_xml, owner_email_from_excel)
            else:
                '''test string trim & multi-email
                '''
                owner_email_from_xml = cfg_parser.getOwnerEmailsOfProject(project)
                self.assertEqual(sorted(owner_email_from_xml), ['bin.yin@citrix.com', 'yang.zhang@citrix.com'])
    
    def test_project_owner_name_parser(self):
        '''get project owner name list from excel & xml, compare them
        '''
        cfg_parser = config_parser.CfgParser(CONFIG_FILE)
        project_list_from_xml = cfg_parser.getProjectList()
        for project in project_list_from_xml:
            if(project.split('-')[0]!='XenDesktop'):
                owner_email_from_xml = cfg_parser.getOwnerEmailsOfProject(project)[0]
                owner_name_from_xml = cfg_parser.getOwnerName(project, owner_email_from_xml)
                                
                wb = load_workbook(PRODUCT_OWNER_LIST_FILE)
                sheet_ranges = wb['Sheet1']
                column_length = int(sheet_ranges.dimensions.split(':')[1].split('C')[1]) 
                for i in range(2, column_length+1):
                    index = str(i)
                    if(project==sheet_ranges['A'+index].value):
                        owner_name_from_excel = sheet_ranges['B'+index].value
                        self.assertEqual(owner_name_from_xml, owner_name_from_excel)
            else:
                self.assertEqual(cfg_parser.getOwnerName(project, 'bin.yin@citrix.com'), 'Bin Yin')
    
    def test_project_threshold_issueCount_parser(self):
        cfg_parser = config_parser.CfgParser(CONFIG_FILE)
        project_list_from_xml = cfg_parser.getProjectList()
        for project in project_list_from_xml:
            threshold_issueCount = cfg_parser.getThresholdIssueCountOfProject(project)
            if(project.split('-')[0]!='XenDesktop'):
                self.assertEqual(threshold_issueCount, 500)
            else:
                self.assertEqual(threshold_issueCount, 123)
    
    def test_project_threshold_issueIncreaseRate_parser(self):
        cfg_parser = config_parser.CfgParser(CONFIG_FILE)
        project_list_from_xml = cfg_parser.getProjectList()
        for project in project_list_from_xml:
            threshold_issueIncreaseRate = cfg_parser.getThresholdIssueIncreaseRateOfProject(project)
            if(project.split('-')[0]!='XenDesktop'):
                self.assertEqual(threshold_issueIncreaseRate, 0.3)
            else:
                self.assertEqual(threshold_issueIncreaseRate, 0.02131234)
                
    def test_project_threshold_issueDecreaseRate_parser(self):
        cfg_parser = config_parser.CfgParser(CONFIG_FILE)
        project_list_from_xml = cfg_parser.getProjectList()
        for project in project_list_from_xml:
            threshold_issueDecreaseRate = cfg_parser.getThresholdIssueDecreaseRateOfProject(project)
            if(project.split('-')[0]!='XenDesktop'):
                self.assertEqual(threshold_issueDecreaseRate, 0.5)
            else:
                self.assertEqual(threshold_issueDecreaseRate, 10)


if __name__ == '__main__':
    unittest.main()